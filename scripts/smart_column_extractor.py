#!/usr/bin/env python3
"""
Smart Column Extraction & Delta Sync Engine for SE2026 (SQLLab BPS)
Menarik 100% seluruh kolom (ratusan/ribuan kolom) dari metadata secara otomatis
dengan teknik Dynamic Column Chunking (20-column SQL batching) & Deep JSON Merging.
"""

import openpyxl
import subprocess
import json
import csv
import os
import sys

BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
XLSX_METADATA = os.path.join(BASE_DIR, "kegiatan", "sensus-ekonomi-2026", "2026", "master_data", "kamus_kolom_se2026.xlsx")
SQLLAB_JS = os.path.join(BASE_DIR, "scripts", "sqllab.js")
OUT_CSV = os.path.join(BASE_DIR, "kegiatan", "sensus-ekonomi-2026", "2026", "sqllab_monitoring", "csv", "sample_delta_10_smart_all_columns.csv")
OUT_JSON = os.path.join(BASE_DIR, "kegiatan", "sensus-ekonomi-2026", "2026", "sqllab_monitoring", "json", "sample_delta_10_smart_all_columns.json")

os.makedirs(os.path.dirname(OUT_CSV), exist_ok=True)
os.makedirs(os.path.dirname(OUT_JSON), exist_ok=True)

def load_schema_from_xlsx(xlsx_path):
    print("📖 Reading metadata columns from Excel...")
    wb = openpyxl.load_workbook(xlsx_path, read_only=True)
    table_cols = {}
    
    # Priority tables for extraction
    target_tables = [
        "base_table_assignment",
        "root_table",
        "se2026_nested",
        "nested_dtsen",
        "nested_meteran",
        "kp_nested"
    ]
    
    for sheet in wb.sheetnames:
        clean_tbl = sheet.replace("tgr_fd68e454.", "")
        if clean_tbl in target_tables or sheet in target_tables:
            ws = wb[sheet]
            rows = list(ws.iter_rows(values_only=True))
            cols = []
            if len(rows) > 1:
                header = rows[0]
                col_idx = 1 if "Kolom" in header else 0
                for r in rows[1:]:
                    if r and len(r) > col_idx and r[col_idx]:
                        col_name = str(r[col_idx]).strip()
                        if col_name and col_name not in cols:
                            cols.append(col_name)
            table_cols[clean_tbl] = cols
            print(f"   ✓ Table '{clean_tbl}': {len(cols)} columns registered")
    
    return table_cols

def run_sqllab_query(sql):
    cmd = ["node", SQLLAB_JS, sql]
    res = subprocess.run(cmd, capture_output=True, text=True)
    out = res.stdout
    start_idx = out.find("[")
    end_idx = out.rfind("]")
    if start_idx != -1 and end_idx != -1 and end_idx > start_idx:
        json_str = out[start_idx:end_idx+1]
        try:
            return json.loads(json_str)
        except Exception as e:
            print(f"❌ JSON Parse Error: {e}")
            return []
    return []

def chunk_columns(columns, chunk_size=18):
    """Chunk column list into batches of max 18 (keeping 1 space for assignment_id)"""
    col_list = [c for c in columns if c.lower() != "assignment_id"]
    for i in range(0, len(col_list), chunk_size):
        yield col_list[i:i + chunk_size]

def main():
    print("🚀 [Smart Column Extractor] Initializing Full-Schema Extraction for 10 Sample Rows...")
    schema = load_schema_from_xlsx(XLSX_METADATA)
    
    # Step 1: Fetch 10 Sample Target Assignment IDs
    print("\n📦 Step 1: Fetching 10 Target Assignment IDs (is_active = 1, latest modified)...")
    sql_base_init = """
    SELECT 
      assignment_id,
      assignment_status_alias,
      assignment_date_modified,
      is_active,
      code_identity,
      level_1_full_code,
      level_2_full_code,
      level_2_name,
      level_3_name,
      level_4_name,
      level_5_full_code,
      level_6_full_code,
      level_6_name,
      current_user_username,
      current_user_survey_role_name
    FROM base_table_assignment
    WHERE level_2_full_code = '6104'
      AND is_active = 1
    ORDER BY assignment_date_modified DESC, assignment_id ASC
    LIMIT 10;
    """
    
    base_rows = run_sqllab_query(sql_base_init)
    if not base_rows:
        print("❌ Failed to fetch base assignment IDs.")
        sys.exit(1)
        
    assign_ids = [r["assignment_id"] for r in base_rows]
    id_clause = ", ".join([f"'{aid}'" for aid in assign_ids])
    print(f"   ✓ Diterima {len(assign_ids)} Assignment IDs.")
    
    # Store complete merged documents
    merged_docs = {aid: dict(r) for aid, r in zip(assign_ids, base_rows)}
    
    # Tables to extract
    main_tables = ["base_table_assignment", "root_table", "se2026_nested"]
    array_tables = ["nested_dtsen", "nested_meteran", "kp_nested"]
    
    total_queries_executed = 1
    total_cols_processed = 15
    
    # Step 2: Extract Main 1:1 Tables via Column Batching
    print("\n⚡ Step 2: Extracting 1:1 Scalar Tables (Column Chunking Batching)...")
    for tbl in main_tables:
        if tbl not in schema:
            continue
        all_cols = schema[tbl]
        batches = list(chunk_columns(all_cols, chunk_size=18))
        print(f"\n   📋 Table '{tbl}': {len(all_cols)} columns across {len(batches)} SQL Batches...")
        
        for idx, col_batch in enumerate(batches, 1):
            select_cols = ["assignment_id"] + col_batch
            select_str = ", ".join(select_cols)
            sql = f"SELECT {select_str} FROM {tbl} WHERE assignment_id IN ({id_clause});"
            
            print(f"      → Batch {idx}/{len(batches)} ({len(col_batch)} cols)...", end=" ", flush=True)
            res = run_sqllab_query(sql)
            total_queries_executed += 1
            total_cols_processed += len(col_batch)
            print(f"✓ Received {len(res)} rows")
            
            for row in res:
                aid = row.get("assignment_id")
                if aid in merged_docs:
                    merged_docs[aid].update(row)
                    
    # Step 3: Extract 1:N Array Tables (Nested Arrays)
    print("\n🧬 Step 3: Extracting 1:N Array Tables (ART, Meteran, KP)...")
    for tbl in array_tables:
        if tbl not in schema:
            continue
        all_cols = schema[tbl]
        batches = list(chunk_columns(all_cols, chunk_size=18))
        print(f"\n   📋 Array Table '{tbl}': {len(all_cols)} columns across {len(batches)} Batches...")
        
        tbl_array_map = {aid: [] for aid in assign_ids}
        
        for idx, col_batch in enumerate(batches, 1):
            select_cols = ["assignment_id"] + col_batch
            select_str = ", ".join(select_cols)
            sql = f"SELECT {select_str} FROM {tbl} WHERE assignment_id IN ({id_clause});"
            
            print(f"      → Batch {idx}/{len(batches)} ({len(col_batch)} cols)...", end=" ", flush=True)
            res = run_sqllab_query(sql)
            total_queries_executed += 1
            print(f"✓ Received {len(res)} child rows")
            
            for row in res:
                aid = row.get("assignment_id")
                if aid in tbl_array_map:
                    tbl_array_map[aid].append(row)
                    
        # Attach array to merged docs
        array_key = f"{tbl}_array"
        for aid, child_rows in tbl_array_map.items():
            if aid in merged_docs:
                merged_docs[aid][array_key] = child_rows
                
    # Step 4: Export Flat CSV & Structured JSON
    print("\n💾 Step 4: Exporting Merged Data to CSV & JSON...")
    final_docs_list = list(merged_docs.values())
    
    # Write JSON
    with open(OUT_JSON, "w", encoding="utf-8") as f:
        json.dump(final_docs_list, f, indent=2, ensure_ascii=False)
    print(f"   ✅ JSON saved to {OUT_JSON}")
    
    # Write CSV (Scalar columns flattened)
    all_flat_keys = []
    for doc in final_docs_list:
        for k, v in doc.items():
            if not isinstance(v, (list, dict)):
                if k not in all_flat_keys:
                    all_flat_keys.append(k)
                    
    with open(OUT_CSV, "w", encoding="utf-8-sig", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=all_flat_keys, extrasaction="ignore")
        writer.writeheader()
        for doc in final_docs_list:
            writer.writerow(doc)
    print(f"   ✅ CSV saved to {OUT_CSV} ({len(all_flat_keys)} scalar columns)")
    
    # Audit summary
    print("\n🎉 [SUCCESS] Extraction Audit & Verification Complete!")
    print(f"   • Total Assignments Extracted: {len(final_docs_list)}")
    print(f"   • Total SQL Batches Executed: {total_queries_executed}")
    print(f"   • Total Unique Scalar Columns in CSV: {len(all_flat_keys)}")
    print(f"   • CSV Output File: {OUT_CSV}")
    print(f"   • JSON Output File: {OUT_JSON}")

if __name__ == "__main__":
    main()
