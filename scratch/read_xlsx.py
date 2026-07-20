import openpyxl

wb = openpyxl.load_workbook("kegiatan/kecamatan-dalam-angka/2026/daftar-perubahan-template-dda-kecamatan-2026.xlsx")
print("Sheets:", wb.sheetnames)

for sheetname in wb.sheetnames:
    sheet = wb[sheetname]
    print(f"\n--- Sheet: {sheetname} ---")
    rows = list(sheet.iter_rows(values_only=True))
    for r in rows[:10]:  # print first 10 rows to inspect
        print(r)
