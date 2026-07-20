import openpyxl

wb = openpyxl.load_workbook("kegiatan/kecamatan-dalam-angka/2026/daftar-perubahan-template-dda-kecamatan-2026.xlsx")

def clean_val(v):
    if v is None:
        return ""
    s = str(v).strip()
    return s

# Process Daftar Perubahan 2026
sheet1 = wb['Daftar Perubahan 2026']
out = []
out.append("# Daftar Perubahan Template DDA Kecamatan 2026\n")
out.append("Dokumen ini berisi daftar perubahan nomor dan judul tabel dari template DDA Kecamatan 2025 ke template 2026.\n")

rows = list(sheet1.iter_rows(values_only=True))
# Clean up empty rows
rows = [r for r in rows if any(x is not None for x in r)]

# Header
out.append("| No Tabel 2025 | Judul Tabel 2025 | No Tabel 2026 | Judul Tabel 2026 | Perubahan (Sebelum) | Perubahan (Sesudah) | Keterangan |")
out.append("| :---: | :--- | :---: | :--- | :--- | :--- | :--- |")

for i, r in enumerate(rows):
    if i < 3: # skip titles and header lines
        continue
    # Columns we want:
    # 0: No Lama, 1: Judul Lama, 3: No Baru, 4: Judul Baru, 6: Sebelum, 7: Sesudah, 8: Keterangan
    # Let's handle index safety:
    no_lama = clean_val(r[0]) if len(r) > 0 else ""
    judul_lama = clean_val(r[1]) if len(r) > 1 else ""
    no_baru = clean_val(r[3]) if len(r) > 3 else ""
    judul_baru = clean_val(r[4]) if len(r) > 4 else ""
    sebelum = clean_val(r[6]) if len(r) > 6 else ""
    sesudah = clean_val(r[7]) if len(r) > 7 else ""
    keterangan = clean_val(r[8]) if len(r) > 8 else ""
    
    # If all empty, skip
    if not (no_lama or judul_lama or no_baru or judul_baru or sebelum or sesudah or keterangan):
        continue
        
    out.append(f"| {no_lama} | {judul_lama} | {no_baru} | {judul_baru} | {sebelum} | {sesudah} | {keterangan} |")

# Process List tabel 2026
sheet2 = wb['List tabel 2026']
out.append("\n---\n")
out.append("# List Seluruh Tabel Template KCDA 2026\n")
out.append("Berikut adalah daftar seluruh tabel yang ada pada template Kecamatan Dalam Angka 2026:\n")
out.append("| No Tabel | Judul Tabel | Perkiraan Tahun Data | Tabel Wajib | Sumber Data | Keterangan |")
out.append("| :---: | :--- | :---: | :---: | :--- | :--- |")

rows2 = list(sheet2.iter_rows(values_only=True))
rows2 = [r for r in rows2 if any(x is not None for x in r)]

for i, r in enumerate(rows2):
    if i < 4: # skip headers
        continue
    no_tabel = clean_val(r[0]) if len(r) > 0 else ""
    judul = clean_val(r[1]) if len(r) > 1 else ""
    tahun = clean_val(r[2]) if len(r) > 2 else ""
    wajib = clean_val(r[3]) if len(r) > 3 else ""
    sumber = clean_val(r[4]) if len(r) > 4 else ""
    ket = clean_val(r[5]) if len(r) > 5 else ""
    
    if not (no_tabel or judul or tahun or wajib or sumber or ket):
        continue
        
    out.append(f"| {no_tabel} | {judul} | {tahun} | {wajib} | {sumber} | {ket} |")

with open("kegiatan/kecamatan-dalam-angka/2026/daftar-perubahan-dda-2026.md", "w", encoding="utf-8") as f:
    f.write("\n".join(out))

print("Markdown conversion complete.")
